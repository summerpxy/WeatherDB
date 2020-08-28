# 操作数据库
import sqlite3

from openpyxl import load_workbook


def get_location_info_by_xlsx():
    wb = load_workbook("AMap_adcode_citycode.xlsx")
    ws = wb["Sheet1"]
    result = []
    for row in ws.iter_rows(min_col=1, min_row=2):
        tmp = []
        for cell in row:
            tmp.append(cell.value)
        result.append(tuple(tmp))

    return result


if __name__ == "__main__":
    print(get_location_info_by_xlsx())
